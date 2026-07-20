"""
Genera i codici ref mancanti per una lista contatti — Rilievo Contract.

Accetta sia CSV che XLSX in input (rilevato dall'estensione — nessuna
opzione da passare). Per XLSX legge TUTTI i fogli tranne uno chiamato
"Riepilogo" e li tratta come un'unica lista di contatti ai fini del
controllo duplicati/collisioni: se il file ha una tab "Architetti" e una
"Agenzie immobiliari", un ref generato per un'agenzia viene comunque
controllato contro quelli degli architetti (e viceversa) — cosi' due
pubblici diversi non possono mai finire con lo stesso codice per errore.
L'input non viene mai toccato: la lista contatti e' l'unica copia della
mappatura ref -> persona (i ref gia' spediti nelle email NON sono
ricostruibili da nessuna formula, vedi sotto), quindi lo script scrive
sempre su un file nuovo.

L'output rispetta il formato dell'input: XLSX in -> XLSX out (stessa
struttura a fogli, stessa formattazione), CSV in -> CSV out. Con --out
puoi comunque forzare l'altro formato (es. leggere un XLSX e scrivere un
CSV) cambiando semplicemente l'estensione del percorso passato.

FORMULA DEI REF NUOVI (aggiornata il 20/07/2026):
I ref sono codici CASUALI a 5 caratteri esadecimali (es. "a1f1c"), non
derivati da nessuna formula (non piu' md5(email) come in una versione
precedente di questo script) — scelta per restare coerenti con i 42 ref
storici, anch'essi casuali e non rigenerabili (vedi
_reference/PLAYBOOK_NUOVO_PUBBLICO.md, sezione "Ref code non
deterministici"). Implicazione pratica: il file con la mappatura
ref -> contatto e' l'UNICA copia di quell'informazione al mondo. Tienilo
al sicuro — backup vero, non solo locale — perche' se si perde non c'e'
calcolo che lo ricostruisca (e' gia' successo, vedi STATO_PROGETTO...).

USO:
    python genera_ref.py nuovi_contatti.csv
    python genera_ref.py Lista_Contatti_Rilievo_Contract.xlsx
    python genera_ref.py lista.xlsx --out lista_aggiornata.xlsx
"""

import argparse
import csv
import re
import secrets
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIBILE = True
except ImportError:
    OPENPYXL_DISPONIBILE = False

FORMATO_REF = re.compile(r"^[0-9a-f]{5}$")
FOGLIO_DA_IGNORARE = "riepilogo"

# Intestazioni "belle" (come in Lista_Contatti_Rilievo_Contract.xlsx, es.
# "Nome", "Città", "Link presentazione") -> nome colonna canonico usato
# internamente dallo script (minuscolo, senza accenti, come nel CSV).
ALIAS_COLONNE = {
    "id": "id",
    "nome": "nome",
    "città": "citta",
    "citta": "citta",
    "email": "email",
    "categoria": "categoria",
    "ref": "ref",
    "link": "link",
    "link presentazione": "link",
    "tipo": "tipo",
}


def canonicalizza_intestazione(testo):
    """Normalizza un'intestazione di colonna (qualunque maiuscole/accenti
    usati per la leggibilita' nell'Excel) al nome canonico usato dallo
    script. Colonne non riconosciute passano invariate (minuscole)."""
    chiave = (testo or "").strip().lower()
    return ALIAS_COLONNE.get(chiave, chiave)

# --- stile Excel, coerente con Lista_Contatti_Rilievo_Contract.xlsx ---
COLORE_HEADER_BG = "1C1C1A"
COLORE_HEADER_TESTO = "F8F5EF"
COLORE_ALT_RIGA = "F8F5EF"
COLORE_BORDO = "D9D6CE"
COLORE_LINK = "0563C1"


def normalizza_email(email):
    return (email or "").strip().lower()


def genera_codice_casuale(occupati):
    """5 caratteri esadecimali casuali, ritenta finche' non ne trova uno libero."""
    while True:
        candidato = secrets.token_hex(3)[:5]
        if candidato not in occupati:
            return candidato


# ------------------------- lettura -------------------------

def leggi_csv(percorso):
    with open(percorso, newline="", encoding="utf-8-sig") as f:
        lettore = csv.DictReader(f)
        colonne = list(lettore.fieldnames or [])
        righe = list(lettore)
    for riga in righe:
        riga["_foglio"] = None
    return {"formato": "csv", "fogli": [None], "colonne_per_foglio": {None: colonne}, "righe": righe}


def leggi_xlsx(percorso):
    if not OPENPYXL_DISPONIBILE:
        sys.exit("Serve 'openpyxl' per leggere file .xlsx: pip install openpyxl --break-system-packages")

    wb = load_workbook(percorso, data_only=True)
    fogli = [nome for nome in wb.sheetnames if nome.strip().lower() != FOGLIO_DA_IGNORARE]
    if not fogli:
        sys.exit(f"Nessun foglio dati in {percorso} (solo '{wb.sheetnames}', tutti esclusi come riepilogo).")

    righe = []
    colonne_per_foglio = {}
    for nome_foglio in fogli:
        ws = wb[nome_foglio]
        righe_ws = list(ws.iter_rows(values_only=True))
        if not righe_ws:
            continue
        intestazione = [canonicalizza_intestazione(c) for c in righe_ws[0]]
        colonne_per_foglio[nome_foglio] = intestazione
        for riga_valori in righe_ws[1:]:
            if all(v is None or str(v).strip() == "" for v in riga_valori):
                continue
            riga = {}
            for indice, nome_colonna in enumerate(intestazione):
                if not nome_colonna:
                    continue
                valore = riga_valori[indice] if indice < len(riga_valori) else None
                riga[nome_colonna] = "" if valore is None else str(valore).strip()
            riga["_foglio"] = nome_foglio
            righe.append(riga)

    return {"formato": "xlsx", "fogli": fogli, "colonne_per_foglio": colonne_per_foglio, "righe": righe}


def leggi_contatti(percorso):
    estensione = Path(percorso).suffix.lower()
    if estensione == ".csv":
        return leggi_csv(percorso)
    if estensione in (".xlsx", ".xlsm"):
        return leggi_xlsx(percorso)
    sys.exit(f"Formato non supportato: '{estensione}' (atteso .csv o .xlsx).")


# ------------------------- scrittura -------------------------

def scrivi_csv(percorso, dati):
    colonne = dati["colonne_per_foglio"][None]
    if "ref" not in colonne:
        colonne = colonne + ["ref"]
    with open(percorso, "w", newline="", encoding="utf-8") as f:
        scrittore = csv.DictWriter(f, fieldnames=colonne, extrasaction="ignore")
        scrittore.writeheader()
        scrittore.writerows(dati["righe"])


def scrivi_xlsx(percorso, dati):
    if not OPENPYXL_DISPONIBILE:
        sys.exit("Serve 'openpyxl' per scrivere file .xlsx: pip install openpyxl --break-system-packages")

    wb = Workbook()
    wb.remove(wb.active)

    # Foglio Riepilogo, sempre rigenerato (mai copiato dall'input): la
    # lettura ignora deliberatamente un foglio chiamato "Riepilogo" (non e'
    # dati di contatti), quindi va ricreato qui o si perderebbe ad ogni
    # giro CSV/XLSX -> genera_ref.py -> XLSX (bug reale, corretto il
    # 20/07/2026: un primo giro aveva fatto sparire il foglio Riepilogo
    # da Lista_Contatti_Rilievo_Contract.xlsx).
    ws_riepilogo = wb.create_sheet("Riepilogo")
    ws_riepilogo.sheet_view.showGridLines = False
    ws_riepilogo.column_dimensions["A"].width = 34
    ws_riepilogo.column_dimensions["B"].width = 14
    titolo = ws_riepilogo.cell(row=1, column=1, value="Lista contatti — Rilievo Contract")
    titolo.font = Font(name="Arial", size=14, bold=True, color="1C1C1A")
    ws_riepilogo.merge_cells("A1:B1")
    sotto = ws_riepilogo.cell(row=2, column=1, value="Mappatura ref -> contatto, tutti i canali")
    sotto.font = Font(name="Arial", size=10, italic=True, color="7A7A75")
    ws_riepilogo.merge_cells("A2:B2")

    header_fill = PatternFill(start_color=COLORE_HEADER_BG, end_color=COLORE_HEADER_BG, fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color=COLORE_HEADER_TESTO)
    body_font = Font(name="Arial", size=10, color="1C1C1A")
    link_font = Font(name="Arial", size=10, color=COLORE_LINK, underline="single")
    alt_fill = PatternFill(start_color=COLORE_ALT_RIGA, end_color=COLORE_ALT_RIGA, fill_type="solid")
    thin = Side(style="thin", color=COLORE_BORDO)
    bordo = Border(left=thin, right=thin, top=thin, bottom=thin)

    righe_riepilogo = [("Tab", "Contatti")]
    for nome_foglio in dati["fogli"]:
        righe_riepilogo.append((nome_foglio, len([r for r in dati["righe"] if r["_foglio"] == nome_foglio])))
    righe_riepilogo.append(("Totale", len(dati["righe"])))

    start_row = 4
    for i, (etichetta, valore) in enumerate(righe_riepilogo):
        r = start_row + i
        c1 = ws_riepilogo.cell(row=r, column=1, value=etichetta)
        c2 = ws_riepilogo.cell(row=r, column=2, value=valore)
        if i == 0:
            c1.font = header_font
            c2.font = header_font
            c1.fill = header_fill
            c2.fill = header_fill
        elif i == len(righe_riepilogo) - 1:
            c1.font = Font(name="Arial", size=10, bold=True, color="1C1C1A")
            c2.font = Font(name="Arial", size=10, bold=True, color="1C1C1A")
        else:
            c1.font = body_font
            c2.font = body_font
        c1.border = bordo
        c2.border = bordo

    for nome_foglio in dati["fogli"]:
        colonne = dati["colonne_per_foglio"][nome_foglio]
        if "ref" not in colonne:
            colonne = colonne + ["ref"]
        righe_foglio = [r for r in dati["righe"] if r["_foglio"] == nome_foglio]

        ws = wb.create_sheet(nome_foglio)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        for col_idx, nome_colonna in enumerate(colonne, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome_colonna.capitalize())
            cell.font = header_font
            cell.fill = header_fill
            cell.border = bordo
            ws.column_dimensions[get_column_letter(col_idx)].width = 30 if nome_colonna != "link" else 52

        for r_idx, riga in enumerate(righe_foglio, start=2):
            for c_idx, nome_colonna in enumerate(colonne, start=1):
                valore = riga.get(nome_colonna, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=valore)
                cell.border = bordo
                cell.alignment = Alignment(vertical="center")
                if nome_colonna == "link" and valore:
                    cell.font = link_font
                    cell.hyperlink = valore
                else:
                    cell.font = body_font
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
        if righe_foglio:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(colonne))}{len(righe_foglio) + 1}"

    wb.save(percorso)


def scrivi_contatti(percorso, dati):
    estensione = Path(percorso).suffix.lower()
    if estensione == ".csv":
        scrivi_csv(percorso, dati)
    elif estensione in (".xlsx", ".xlsm"):
        scrivi_xlsx(percorso, dati)
    else:
        sys.exit(f"Formato di output non supportato: '{estensione}' (atteso .csv o .xlsx).")


# ------------------------- logica principale -------------------------

def main():
    parser = argparse.ArgumentParser(description="Popola i ref mancanti in una lista contatti (CSV o XLSX)")
    parser.add_argument("file_contatti", help="CSV o XLSX con almeno le colonne nome,email (ref opzionale)")
    parser.add_argument("--out", help="File di output (default: <input>_con_ref.<stessa estensione>)")
    args = parser.parse_args()

    percorso_input = Path(args.file_contatti)
    if not percorso_input.exists():
        sys.exit(f"File non trovato: {args.file_contatti}")

    if args.out:
        percorso_output = Path(args.out)
    else:
        percorso_output = percorso_input.with_name(percorso_input.stem + "_con_ref" + percorso_input.suffix)

    if percorso_output.resolve() == percorso_input.resolve():
        sys.exit("Il file di output coincide con l'input: scegli un --out diverso, l'input non va mai sovrascritto.")

    dati = leggi_contatti(percorso_input)
    righe = dati["righe"]

    if not righe:
        sys.exit("Nessuna riga di contatti trovata.")
    for nome_foglio, colonne in dati["colonne_per_foglio"].items():
        for colonna in ["nome", "email"]:
            if colonna not in colonne:
                etichetta = f"foglio '{nome_foglio}'" if nome_foglio else "il CSV"
                sys.exit(f"Colonna '{colonna}' mancante in {etichetta} (richieste: nome, email).")
    for riga in righe:
        if "ref" not in riga:
            riga["ref"] = ""

    avvisi = []

    # 1. Censimento email duplicate (stessa persona, non deve avere due ref) —
    #    su TUTTI i fogli insieme, cosi' la stessa persona non riceve due ref
    #    diversi se compare per errore sia tra gli architetti sia tra le agenzie.
    righe_per_email = {}
    for indice, riga in enumerate(righe):
        email = normalizza_email(riga.get("email"))
        righe_per_email.setdefault(email, []).append(indice)
    for email, indici in righe_per_email.items():
        if len(indici) > 1:
            luoghi = [f"riga {i + 2}" + (f" ({righe[i]['_foglio']})" if righe[i]["_foglio"] else "") for i in indici]
            avvisi.append(
                f"EMAIL DUPLICATA: '{email}' compare {len(indici)} volte ({', '.join(luoghi)}) — "
                "stessa persona, ricevera' lo stesso ref. Valuta se e' un doppione da rimuovere."
            )

    # 2. Censimento ref gia' assegnati (mai toccati) + controllo formato —
    #    anche qui su TUTTI i fogli insieme: e' questo che garantisce che un
    #    ref nuovo per le agenzie non collida con uno gia' usato dagli
    #    architetti, e viceversa.
    ref_occupati = set()
    for indice, riga in enumerate(righe):
        ref_esistente = (riga.get("ref") or "").strip()
        riga["ref"] = ref_esistente
        if ref_esistente == "":
            continue
        if not FORMATO_REF.match(ref_esistente):
            luogo = f"riga {indice + 2}" + (f" ({riga['_foglio']})" if riga["_foglio"] else "")
            avvisi.append(
                f"FORMATO ANOMALO: {luogo} ha ref '{ref_esistente}' (atteso: 5 caratteri esadecimali) — "
                "lasciato invariato, ma verifica che sia voluto."
            )
        if ref_esistente in ref_occupati:
            avvisi.append(
                f"REF DUPLICATO GIA' PRESENTE: '{ref_esistente}' compare piu' volte nell'input — "
                "due contatti diversi con lo stesso codice sono indistinguibili in GA4, correggere a mano."
            )
        ref_occupati.add(ref_esistente)

    # 3. Generazione casuale per le righe senza ref, una email per volta cosi'
    #    i duplicati della stessa email ricevono lo stesso codice.
    ref_per_email = {}
    for email, indici in righe_per_email.items():
        for indice in indici:
            if righe[indice]["ref"] != "":
                ref_per_email[email] = righe[indice]["ref"]
                break

    generati = 0
    ereditati = 0
    for email, indici in righe_per_email.items():
        for indice in indici:
            if righe[indice]["ref"] != "":
                continue
            if email in ref_per_email:
                righe[indice]["ref"] = ref_per_email[email]
                ereditati += 1
                continue

            codice = genera_codice_casuale(ref_occupati)
            righe[indice]["ref"] = codice
            ref_occupati.add(codice)
            ref_per_email[email] = codice
            generati += 1

    scrivi_contatti(percorso_output, dati)

    for avviso in avvisi:
        print("ATTENZIONE —", avviso)
    print(f"Fatto. {len(righe)} righe scritte in: {percorso_output}")
    if len(dati["fogli"]) > 1:
        print(f"Fogli letti insieme (stesso controllo duplicati/collisioni): {', '.join(dati['fogli'])}")
    print(f"Ref generati (casuali): {generati} | ereditati da email duplicata: {ereditati} | "
          f"gia' presenti e non toccati: {len(righe) - generati - ereditati}")


if __name__ == "__main__":
    main()
