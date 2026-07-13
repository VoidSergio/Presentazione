"""
GA4 Report Builder — Rilievo Contract

Legge un export CSV da un'Esplorazione GA4 (scheda "Profilo per contatto",
formato libero, con o senza Confronti di segmenti) e produce un file
Excel a tre fogli: Riepilogo (una riga per contatto, leggibile da
chiunque), Dettaglio eventi (conteggio per singolo evento) e Traffico
anonimo (traffico senza ref, aggregato per evento).

Opzionalmente unisce i dati con un file di mappatura ref -> nome studio
(NON incluso in questo script, va fornito da te in locale — non caricarlo
mai in una chat o in un repository condiviso, contiene dati sensibili).

USO:
    python ga4_report_builder.py export.csv
    python ga4_report_builder.py export.csv --contacts contatti.csv
    python ga4_report_builder.py export.csv --contacts contatti.csv --out report.xlsx

Il file --contacts, se fornito, deve avere almeno le colonne:
    ref,studio      (oppure ref,nome — 'nome' viene accettato come alias
                     di 'studio', cosi' si puo' usare direttamente la
                     lista contatti gia' esistente senza rinominare nulla)
(altre colonne extra come email, città, ecc. sono ignorate ma tollerate)

STRUTTURA ATTESA DELL'EXPORT (scheda "Profilo per contatto — Produzione")
Dimensioni riconosciute per NOME (ordine e presenza non garantiti):
    ref (obbligatoria), Nome evento (obbligatoria), Nome host (opzionale),
    Categoria del dispositivo (opzionale), slide_number (opzionale),
    type (opzionale)
Metriche: Conteggio eventi (obbligatoria), seconds_spent (opzionale)
Ogni dimensione/metrica opzionale mancante disattiva la relativa colonna
derivata nel report con un messaggio a video, senza mai far fallire lo
script — vedi README per la tabella completa obbligatorie/opzionali.

OUTPUT:
    --out con estensione .xlsx (default) -> tre fogli:
        "Riepilogo"        — una riga per contatto identificabile
        "Dettaglio eventi"  — stesso profilo, conteggio per singolo evento
        "Traffico anonimo"  — eventi senza ref, aggregati (non per persona)
    --out con estensione .csv -> solo il foglio Riepilogo, in CSV piatto
"""

import argparse
import csv
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXCLUDE_HOST_SUBSTRINGS = ["localhost", "netlify.app"]

# Colonne che sono SEMPRE dimensioni (testo), mai metriche: non vanno mai
# convertite in numeri, anche se qualche valore sembra numerico. Il caso
# concreto: un ref come "53775" e' composto di sole cifre, ma resta un
# codice, non un numero — convertirlo trasformerebbe gli altri ref
# alfanumerici in valori mancanti, collassando contatti diversi.
DIMENSION_COLUMN_SUBSTRINGS = [
    "ref", "nome evento", "event name", "nome host", "hostname", "segmento",
    "segment", "categoria del dispositivo", "device category", "slide_number", "type",
]

# Valori GA4 che indicano "informazione assente" — sia per il ref (visita
# senza ?ref=) sia per qualunque altra dimensione (es. type non ancora
# popolato su eventi antecedenti alla creazione della dimensione).
NOT_SET_VALUES = ["(not set)", "(non impostato)"]

# I 7 eventi custom inviati dal sito (src/utils/analytics.js), nell'ordine
# di priorita' usato in GA4_PARAMETRI_TRACCIAMENTO.md — usato per ordinare
# le colonne del foglio Dettaglio, cosi' i piu' rilevanti stanno a sinistra.
EVENTI_CUSTOM_ORDINE = [
    "contact_click", "full_presentation_viewed", "slide_view",
    "slide_time_spent", "carousel_swipe", "nav_dot_click", "scroll_hint_outcome",
]

# Mappatura fissa numero -> nome slide, ordine post-riordino dell'08/07/2026
# (vedi src/hooks/useSlideNavigation.js, NOMI_SLIDE).
NOMI_SLIDE_PER_NUMERO = {
    1: "Cover", 2: "Lavori scelti", 3: "Cosa facciamo", 4: "Clienti",
    5: "Come lavoriamo", 6: "Chi siamo", 7: "Contatti",
}

TRADUZIONE_CANALE_CONTATTO = {"phone": "Telefono", "email": "Email"}

ETICHETTA_REF_NON_TROVATO = "Ref presente ma non in lista contatti (verificare)"


def _is_dimension_column(column_name: str) -> bool:
    lowered = column_name.strip().lower()
    for substring in DIMENSION_COLUMN_SUBSTRINGS:
        if substring in lowered:
            return True
    return False


def _e_segmento_totale(nome_segmento: str) -> bool:
    """Un gruppo di colonne segmentato chiamato 'Totali'/'Totale' non e'
    un secondo pubblico da sommare al primo: nell'export reale duplica
    esattamente gli stessi numeri del segmento applicato (es. 'Contatti
    reali'), e sommarlo raddoppierebbe ogni conteggio. Riconosciuto per
    nome, non per posizione."""
    return "total" in nome_segmento.strip().lower()


def _colonna_ha_segmento_totale(nome_colonna: str) -> bool:
    if " — " in nome_colonna:
        segmento = nome_colonna.split(" — ", 1)[0]
        return _e_segmento_totale(segmento)
    return False


def _trova_colonna_esatta(df: pd.DataFrame, nome: str):
    """Cerca una colonna per nome esatto (case-insensitive, spazi
    ignorati). Usato per dimensioni con un nome breve/generico (es.
    'type'), dove una ricerca per sottostringa rischierebbe falsi
    positivi su altre colonne."""
    nome_pulito = nome.strip().lower()
    for colonna in df.columns:
        if colonna.strip().lower() == nome_pulito:
            return colonna
    return None


def _trova_colonne_metrica(df: pd.DataFrame, *substrings: str):
    """Trova le colonne metrica corrispondenti a una o piu' sottostringhe
    (es. 'conteggio eventi', 'event count'), escludendo i gruppi
    segmentati 'Totali' quando esiste almeno un'alternativa — vedi
    _e_segmento_totale."""
    candidate = [
        c for c in df.columns
        if any(s in c.lower() for s in substrings)
    ]
    non_totale = [c for c in candidate if not _colonna_ha_segmento_totale(c)]
    if non_totale:
        return non_totale
    return candidate


def _pulisci_non_impostato(serie: pd.Series) -> pd.Series:
    """Sostituisce '(not set)'/'(non impostato)'/cella vuota con NA: per
    una dimensione come 'type' o 'Categoria del dispositivo', questi
    valori significano 'informazione assente per QUESTA riga' (es. evento
    antecedente alla creazione della dimensione), non un valore
    categorico vero da mostrare com'e' nel report."""
    pulita = serie.astype(str).str.strip()
    valori_mancanti = set(v.lower() for v in NOT_SET_VALUES)
    maschera_mancante = pulita.str.lower().isin(valori_mancanti) | (pulita == "")
    return pulita.mask(maschera_mancante)


def parse_ga4_export(path: str) -> pd.DataFrame:
    """Legge un export GA4 'Formato libero', gestendo:
    - righe di commento iniziali (#...)
    - riga opzionale 'Segmento' con nomi dei confronti (uno o piu' gruppi)
    - riga di intestazione colonne
    - riga 'Totale complessivo' da scartare
    Restituisce un DataFrame con colonne pulite.
    """
    # utf-8-sig: gli export GA4 spesso hanno il BOM iniziale, che con
    # "utf-8" semplice finirebbe dentro il nome della prima colonna.
    with open(path, newline="", encoding="utf-8-sig") as f:
        raw_lines = [line for line in csv.reader(f)]

    # Rimuovi righe di commento (# ...) e righe completamente vuote
    lines = [
        row for row in raw_lines
        if not (row and row[0].startswith("#")) and any(cell.strip() for cell in row)
    ]

    if not lines:
        raise ValueError("Nessun dato trovato nel file: controlla che sia un export GA4 valido.")

    idx = 0
    segment_row = None

    # Riga 'Segmento' opzionale: presente solo se hai usato Confronti di
    # segmenti. Puo' avere uno o piu' gruppi (es. solo "Contatti reali",
    # oppure "Contatti reali" + "Totali").
    if "Segmento" in lines[idx]:
        segment_row = lines[idx]
        idx += 1

    header_row = lines[idx]
    idx += 1

    # Costruisci i nomi di colonna finali. La cella che contiene
    # letteralmente il testo "Segmento" e' l'etichetta della riga stessa
    # (compare una volta, subito prima dei nomi di segmento veri), NON un
    # nome di segmento da applicare come prefisso: va ignorata, altrimenti
    # la dimensione che le sta sopra (es. "type") verrebbe rinominata in
    # "Segmento — type" e la ricerca per nome esatto non la troverebbe piu'.
    columns = []
    for i, name in enumerate(header_row):
        valore_segmento = segment_row[i].strip() if segment_row and i < len(segment_row) else ""
        if valore_segmento and valore_segmento.lower() != "segmento":
            columns.append(f"{valore_segmento} — {name.strip()}")
        else:
            columns.append(name.strip())

    data_rows = []
    for row in lines[idx:]:
        if any("totale complessivo" in cell.lower() for cell in row):
            continue  # riga di riepilogo totale, non è un dato reale
        if len(row) < len(columns):
            row = row + [""] * (len(columns) - len(row))
        data_rows.append(row[: len(columns)])

    df = pd.DataFrame(data_rows, columns=columns)

    # Converte in numerica solo una colonna che (a) non e' una dimensione
    # nota e (b) ha TUTTI i valori non vuoti convertibili. La vecchia
    # regola "almeno un valore converte" corrompeva la colonna ref quando
    # conteneva codici di sole cifre insieme a codici alfanumerici; la
    # stessa cosa succederebbe a slide_number, che mischia "(not set)" a
    # numeri veri — quella colonna resta testo qui e viene convertita
    # punto per punto in costruisci_ultima_slide, dove "(not set)" diventa
    # correttamente "valore mancante" invece di essere scartato dall'intera
    # colonna.
    for col in df.columns:
        if _is_dimension_column(col):
            continue
        non_empty = df[col].astype(str).str.strip() != ""
        if non_empty.sum() == 0:
            continue
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted[non_empty].notna().all():
            df[col] = converted.fillna(0)

    return df


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Esclude righe di traffico tecnico (localhost, netlify.app). Se la
    dimensione 'Nome host' non e' nell'export, non fa nulla (opzionale)."""
    host_col = next((c for c in df.columns if "nome host" in c.lower()), None)
    if host_col:
        mask = ~df[host_col].str.lower().str.contains(
            "|".join(EXCLUDE_HOST_SUBSTRINGS), na=False
        )
        df = df[mask]
    return df


def _separa_traffico_con_e_senza_ref(df: pd.DataFrame, ref_col: str):
    """Divide l'export in due parti: righe con un ref identificabile e
    righe senza (testo '(not set)' o cella vuota — stesso significato,
    traffico reale ma non riconducibile a un contatto specifico). Le
    seconde non vengono scartate: finiscono nel foglio 'Traffico
    anonimo', aggregate per evento invece che per persona."""
    df = df.copy()
    df[ref_col] = df[ref_col].astype(str).str.strip()

    senza_ref = df[ref_col].str.lower().isin([v.lower() for v in NOT_SET_VALUES])
    senza_ref = senza_ref | (df[ref_col] == "")

    df_anonimo = df[senza_ref].copy()
    df_con_ref = df[~senza_ref].copy()
    return df_con_ref, df_anonimo


def build_contact_profile(df: pd.DataFrame):
    """Aggrega per ref: conteggio eventi, piu' i campi derivati opzionali
    (ultima slide raggiunta, dispositivo prevalente, canale di contatto,
    tempo totale) quando le dimensioni/metriche relative sono presenti
    nell'export. Restituisce anche il DataFrame del traffico anonimo, per
    il foglio separato."""
    ref_col = _trova_colonna_esatta(df, "ref")
    event_col = next((c for c in df.columns if "nome evento" in c.lower()), None)

    if not ref_col or not event_col:
        raise ValueError(
            "Non trovo le colonne 'ref' e/o 'Nome evento' nell'export. "
            "Verifica di aver incluso entrambe le dimensioni in Esplora."
        )

    df_con_ref, df_anonimo = _separa_traffico_con_e_senza_ref(df, ref_col)

    numero_righe_anonime = len(df_anonimo)
    if numero_righe_anonime > 0:
        print(
            f"{numero_righe_anonime} righe senza ref ('(not set)' o cella vuota) isolate "
            "nel foglio 'Traffico anonimo' — traffico reale ma non riconducibile a un contatto."
        )

    # Metrica eventi: cerca per NOME le colonne "Conteggio eventi",
    # escludendo un eventuale gruppo segmentato "Totali" duplicato (vedi
    # _trova_colonne_metrica). Sommare invece QUALUNQUE colonna numerica
    # mischierebbe metriche diverse (es. "Conteggio eventi" + "Utenti
    # totali" = numero senza senso).
    metric_cols = _trova_colonne_metrica(df_con_ref, "conteggio eventi", "event count")
    if not metric_cols:
        metric_cols = [
            c for c in df_con_ref.columns
            if c not in (ref_col, event_col) and pd.api.types.is_numeric_dtype(df_con_ref[c])
        ]
        if metric_cols:
            print(
                "Attenzione: nessuna colonna 'Conteggio eventi' trovata, uso la somma di: "
                + ", ".join(metric_cols)
                + " — verifica che siano davvero conteggi di eventi."
            )

    df_con_ref["_eventi"] = df_con_ref[metric_cols].sum(axis=1) if metric_cols else 1

    pivot = df_con_ref.pivot_table(
        index=ref_col, columns=event_col, values="_eventi", aggfunc="sum", fill_value=0
    )

    # eventi_slide_view: numero di EVENTI slide_view (le rivisite contano),
    # NON il numero di slide distinte viste — per quello c'e' "ultima
    # slide raggiunta" sotto, quando la dimensione slide_number e' presente.
    pivot["eventi_slide_view"] = pivot.get("slide_view", 0)
    pivot["presentazione_completata"] = pivot.get("full_presentation_viewed", 0) > 0
    pivot["ha_cliccato_contatti"] = pivot.get("contact_click", 0) > 0

    profile = pivot.reset_index().rename(columns={ref_col: "ref"})

    profile = _aggiungi_ultima_slide_raggiunta(profile, df_con_ref, ref_col)
    profile = _aggiungi_dispositivo_prevalente(profile, df_con_ref, ref_col)
    profile = _aggiungi_canale_contatto(profile, df_con_ref, ref_col, event_col)
    profile = _aggiungi_tempo_totale(profile, df_con_ref, ref_col)

    return profile, df_anonimo, event_col


def _aggiungi_ultima_slide_raggiunta(profile, df_con_ref, ref_col):
    """Aggiunge 'ultima_slide_raggiunta' (nome leggibile) se la dimensione
    slide_number e' nell'export. '(not set)' su una riga (evento
    antecedente alla creazione della dimensione) conta come 'mancante per
    quella riga', non entra nel calcolo del massimo — ma non fa fallire
    nulla."""
    colonna = _trova_colonna_esatta(df_con_ref, "slide_number")
    if colonna is None:
        print("Attenzione: dimensione 'slide_number' non presente nell'export — 'Ultima slide raggiunta' omessa dal report.")
        return profile

    numerico = pd.to_numeric(df_con_ref[colonna], errors="coerce")
    massimo_per_ref = numerico.groupby(df_con_ref[ref_col]).max()

    def _nome_slide(numero):
        if pd.isna(numero):
            return pd.NA
        return NOMI_SLIDE_PER_NUMERO.get(int(numero), pd.NA)

    nomi = massimo_per_ref.map(_nome_slide)
    return profile.merge(nomi.rename("ultima_slide_raggiunta"), left_on="ref", right_index=True, how="left")


def _aggiungi_dispositivo_prevalente(profile, df_con_ref, ref_col):
    """Aggiunge 'dispositivo_prevalente' (moda di Categoria del
    dispositivo per ref) se la dimensione e' nell'export."""
    colonna = _trova_colonna_esatta(df_con_ref, "Categoria del dispositivo")
    if colonna is None:
        print("Attenzione: dimensione 'Categoria del dispositivo' non presente nell'export — 'Dispositivo prevalente' omesso dal report.")
        return profile

    pulita = _pulisci_non_impostato(df_con_ref[colonna])

    def _moda(serie):
        valori = serie.dropna()
        if valori.empty:
            return pd.NA
        return valori.mode().iloc[0]

    moda_per_ref = pulita.groupby(df_con_ref[ref_col]).agg(_moda)
    moda_per_ref = moda_per_ref.map(lambda v: v.capitalize() if isinstance(v, str) else v)
    return profile.merge(moda_per_ref.rename("dispositivo_prevalente"), left_on="ref", right_index=True, how="left")


def _aggiungi_canale_contatto(profile, df_con_ref, ref_col, event_col):
    """Aggiunge 'canale_contatto_cliccato': valorizzato solo se esiste
    almeno un evento contact_click con 'type' valorizzato (non
    '(not set)') per quel ref. Se il ref ha cliccato piu' canali diversi,
    li elenca entrambi."""
    colonna = _trova_colonna_esatta(df_con_ref, "type")
    if colonna is None:
        print("Attenzione: dimensione 'type' non presente nell'export — 'Canale di contatto cliccato' omesso dal report.")
        return profile

    tipo_pulito = _pulisci_non_impostato(df_con_ref[colonna])
    maschera_click = df_con_ref[event_col] == "contact_click"

    def _canali(serie):
        valori = sorted(set(serie.dropna()))
        if not valori:
            return pd.NA
        tradotti = [TRADUZIONE_CANALE_CONTATTO.get(v, v) for v in valori]
        return ", ".join(tradotti)

    canali_per_ref = tipo_pulito[maschera_click].groupby(df_con_ref.loc[maschera_click, ref_col]).agg(_canali)
    return profile.merge(canali_per_ref.rename("canale_contatto_cliccato"), left_on="ref", right_index=True, how="left")


def _aggiungi_tempo_totale(profile, df_con_ref, ref_col):
    """Aggiunge 'tempo_totale_sul_sito': somma di seconds_spent per ref,
    se la metrica e' nell'export (esclude un eventuale gruppo 'Totali'
    duplicato, stessa logica di _trova_colonne_metrica)."""
    colonne = _trova_colonne_metrica(df_con_ref, "seconds_spent")
    if not colonne:
        print("Attenzione: metrica 'seconds_spent' non presente nell'export — 'Tempo totale sul sito' omesso dal report.")
        return profile

    numerico = df_con_ref[colonne].apply(pd.to_numeric, errors="coerce").fillna(0)
    df_con_ref = df_con_ref.copy()
    df_con_ref["_secondi_riga"] = numerico.sum(axis=1)
    totale_per_ref = df_con_ref.groupby(ref_col)["_secondi_riga"].sum()
    return profile.merge(totale_per_ref.rename("tempo_totale_sul_sito"), left_on="ref", right_index=True, how="left")


def costruisci_traffico_anonimo(df_anonimo: pd.DataFrame, event_col: str) -> pd.DataFrame:
    """Aggrega il traffico senza ref per evento (e per host, se la
    dimensione e' presente): non una riga per persona (impossibile,
    potrebbero essere piu' visitatori diversi), solo il totale."""
    if df_anonimo.empty:
        return pd.DataFrame(columns=["Nome evento", "Eventi"])

    host_col = next((c for c in df_anonimo.columns if "nome host" in c.lower()), None)
    metric_cols = _trova_colonne_metrica(df_anonimo, "conteggio eventi", "event count")

    df_anonimo = df_anonimo.copy()
    df_anonimo["_eventi"] = df_anonimo[metric_cols].sum(axis=1) if metric_cols else 1

    colonne_raggruppamento = [event_col]
    if host_col:
        colonne_raggruppamento.append(host_col)

    aggregato = df_anonimo.groupby(colonne_raggruppamento)["_eventi"].sum().reset_index()
    rinomina = {event_col: "Nome evento", "_eventi": "Eventi"}
    if host_col:
        rinomina[host_col] = "Nome host"
    aggregato = aggregato.rename(columns=rinomina)

    colonna_ordinamento = ["Nome host", "Nome evento"] if host_col else ["Nome evento"]
    return aggregato.sort_values(colonna_ordinamento, kind="stable").reset_index(drop=True)


def merge_contacts(profile: pd.DataFrame, contacts_path: str) -> pd.DataFrame:
    contacts = pd.read_csv(contacts_path, dtype=str, encoding="utf-8-sig")
    if "ref" not in contacts.columns:
        raise ValueError("Il file contatti deve avere una colonna 'ref'.")

    # 'nome' accettato come alias di 'studio': la lista contatti reale usa
    # gia' quella colonna, inutile pretendere un file rinominato a mano.
    if "studio" not in contacts.columns and "nome" in contacts.columns:
        contacts = contacts.rename(columns={"nome": "studio"})

    contacts["ref"] = contacts["ref"].astype(str).str.strip()

    # Ref duplicati nel file contatti duplicherebbero silenziosamente le
    # righe del report nel merge: tieni la prima occorrenza e avvisa.
    duplicati = contacts["ref"].duplicated()
    if duplicati.sum() > 0:
        print(f"Attenzione: {duplicati.sum()} ref duplicati nel file contatti, tengo la prima occorrenza.")
        contacts = contacts[~duplicati]

    merged = profile.merge(contacts[["ref", "studio"]], on="ref", how="left")

    # Ref presenti nell'export ma assenti dalla lista contatti sono
    # comunque dati validi (una persona reale con un ref proprio, es. una
    # verifica manuale come "check-views" — vedi README): non vengono
    # scartati ne' lasciati vuoti, ricevono un'etichetta esplicita cosi'
    # chi legge il report capisce subito che quella riga non ha un nome
    # studio associato, senza per questo essere rumore da scartare.
    merged["studio"] = merged["studio"].fillna(ETICHETTA_REF_NON_TROVATO)

    cols = ["studio"] + [c for c in merged.columns if c != "studio"]
    return merged[cols]


def costruisci_riepilogo(profile: pd.DataFrame) -> pd.DataFrame:
    """Tabella essenziale, leggibile da chiunque: una riga per contatto,
    le domande che contano ('ha visto tutto?', 'fino a dove?', 'ha
    cliccato un contatto, quale canale?', 'da che dispositivo?', 'quanto
    tempo?'). Le colonne derivate opzionali compaiono solo se erano
    disponibili nell'export. Ordinata alfabeticamente."""
    colonne_base = []
    if "studio" in profile.columns:
        colonne_base.append("studio")
    colonne_base.append("ref")

    colonne_e_nomi = [
        ("presentazione_completata", "Presentazione completata"),
        ("ha_cliccato_contatti", "Ha cliccato un contatto"),
        ("canale_contatto_cliccato", "Canale di contatto cliccato"),
        ("eventi_slide_view", "Slide viste (eventi)"),
        ("ultima_slide_raggiunta", "Ultima slide raggiunta"),
        ("dispositivo_prevalente", "Dispositivo prevalente"),
        ("tempo_totale_sul_sito", "Tempo totale sul sito (secondi)"),
    ]
    colonne_presenti = [c for c, _ in colonne_e_nomi if c in profile.columns]

    riepilogo = profile[colonne_base + colonne_presenti].copy()

    riepilogo["presentazione_completata"] = riepilogo["presentazione_completata"].map({True: "Sì", False: "No"})
    riepilogo["ha_cliccato_contatti"] = riepilogo["ha_cliccato_contatti"].map({True: "Sì", False: "No"})

    mappa_nomi = dict(colonne_e_nomi)
    mappa_nomi.update({"ref": "Ref", "studio": "Studio"})
    riepilogo = riepilogo.rename(columns=mappa_nomi)

    colonna_ordinamento = "Studio" if "Studio" in riepilogo.columns else "Ref"
    return riepilogo.sort_values(colonna_ordinamento, kind="stable").reset_index(drop=True)


def costruisci_dettaglio(profile: pd.DataFrame) -> pd.DataFrame:
    """Tabella completa: una colonna per ogni evento (custom prima, in
    ordine di priorita', poi eventuali eventi automatici GA4 in coda)."""
    colonne_base = []
    if "studio" in profile.columns:
        colonne_base.append("studio")
    colonne_base.append("ref")

    colonne_derivate = {
        "presentazione_completata", "ha_cliccato_contatti", "eventi_slide_view",
        "ultima_slide_raggiunta", "dispositivo_prevalente", "canale_contatto_cliccato",
        "tempo_totale_sul_sito",
    }
    tutte_le_colonne_eventi = [
        c for c in profile.columns
        if c not in colonne_base and c not in colonne_derivate
    ]

    eventi_custom_presenti = [c for c in EVENTI_CUSTOM_ORDINE if c in tutte_le_colonne_eventi]
    eventi_automatici = sorted(c for c in tutte_le_colonne_eventi if c not in eventi_custom_presenti)

    dettaglio = profile[colonne_base + eventi_custom_presenti + eventi_automatici].copy()
    dettaglio = dettaglio.rename(columns={"ref": "Ref", "studio": "Studio"})

    colonna_ordinamento = "Studio" if "Studio" in dettaglio.columns else "Ref"
    dettaglio = dettaglio.sort_values(colonna_ordinamento, kind="stable").reset_index(drop=True)
    return dettaglio, len(eventi_custom_presenti)


def scrivi_xlsx(riepilogo, dettaglio, numero_eventi_custom, traffico_anonimo, percorso_output, righe_anonime):
    with pd.ExcelWriter(percorso_output, engine="openpyxl") as writer:
        riepilogo.to_excel(writer, sheet_name="Riepilogo", index=False, startrow=1)
        dettaglio.to_excel(writer, sheet_name="Dettaglio eventi", index=False, startrow=1)
        traffico_anonimo.to_excel(writer, sheet_name="Traffico anonimo", index=False, startrow=1)

        foglio_riepilogo = writer.sheets["Riepilogo"]
        foglio_riepilogo["A1"] = (
            "Una riga per contatto identificabile (ha un ref proprio, anche se non abbinato "
            "a un nome studio). Il traffico senza ref è nel foglio 'Traffico anonimo'."
        )
        foglio_riepilogo["A1"].font = Font(italic=True, color="55524A")
        _formatta_intestazione_su_riga(foglio_riepilogo, riga_intestazione=2)

        foglio_dettaglio = writer.sheets["Dettaglio eventi"]
        foglio_dettaglio["A1"] = (
            "Conteggio eventi per contatto. Colonne scure = eventi inviati dal sito; "
            "colonne chiare = eventi automatici di GA4 (page_view, scroll, ecc.)."
        )
        foglio_dettaglio["A1"].font = Font(italic=True, color="55524A")
        _formatta_intestazione_su_riga(foglio_dettaglio, riga_intestazione=2, numero_colonne_evidenziate=numero_eventi_custom)

        foglio_anonimo = writer.sheets["Traffico anonimo"]
        foglio_anonimo["A1"] = (
            f"{righe_anonime} righe dell'export senza ref ('(not set)' o cella vuota), aggregate per evento: "
            "traffico reale (visite dirette, non da un link email), non riconducibile a un singolo contatto."
        )
        foglio_anonimo["A1"].font = Font(italic=True, color="55524A")
        _formatta_intestazione_su_riga(foglio_anonimo, riga_intestazione=2)


def _formatta_intestazione_su_riga(worksheet, riga_intestazione, numero_colonne_evidenziate=0):
    intestazione_font = Font(bold=True, color="FFFFFF")
    intestazione_sfondo = PatternFill(start_color="1C1C1A", end_color="1C1C1A", fill_type="solid")
    intestazione_sfondo_automatici = PatternFill(start_color="8A8680", end_color="8A8680", fill_type="solid")

    colonne_iniziali = 2 if worksheet.cell(row=riga_intestazione, column=1).value == "Studio" else 1

    for indice_colonna in range(1, worksheet.max_column + 1):
        cella = worksheet.cell(row=riga_intestazione, column=indice_colonna)
        cella.font = intestazione_font
        if numero_colonne_evidenziate > 0 and indice_colonna > colonne_iniziali + numero_colonne_evidenziate:
            cella.fill = intestazione_sfondo_automatici
        else:
            cella.fill = intestazione_sfondo
        cella.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = f"A{riga_intestazione + 1}"

    for indice_colonna in range(1, worksheet.max_column + 1):
        lunghezza_massima = 0
        for riga in worksheet.iter_rows(min_col=indice_colonna, max_col=indice_colonna, min_row=riga_intestazione):
            for cella in riga:
                if cella.value is not None:
                    lunghezza_massima = max(lunghezza_massima, len(str(cella.value)))
        lettera = get_column_letter(indice_colonna)
        worksheet.column_dimensions[lettera].width = min(max(lunghezza_massima + 3, 10), 55)


def main():
    parser = argparse.ArgumentParser(description="Analizza export GA4 per Rilievo Contract")
    parser.add_argument("export_csv", help="File CSV esportato da GA4 Esplora")
    parser.add_argument("--contacts", help="CSV locale con colonne ref,studio (facoltativo)")
    parser.add_argument("--out", default="report_per_contatto.xlsx",
                        help="File di output: .xlsx (default, 3 fogli) o .csv (solo Riepilogo)")
    args = parser.parse_args()

    if not Path(args.export_csv).exists():
        sys.exit(f"File non trovato: {args.export_csv}")
    if args.contacts and not Path(args.contacts).exists():
        sys.exit(f"File contatti non trovato: {args.contacts}")

    try:
        df = parse_ga4_export(args.export_csv)
        df = clean_dataframe(df)
        profile, df_anonimo, event_col = build_contact_profile(df)
        if args.contacts:
            profile = merge_contacts(profile, args.contacts)
    except ValueError as errore:
        sys.exit(f"Errore: {errore}")

    riepilogo = costruisci_riepilogo(profile)
    dettaglio, numero_eventi_custom = costruisci_dettaglio(profile)
    traffico_anonimo = costruisci_traffico_anonimo(df_anonimo, event_col)

    estensione = Path(args.out).suffix.lower()
    if estensione == ".xlsx":
        scrivi_xlsx(riepilogo, dettaglio, numero_eventi_custom, traffico_anonimo, args.out, len(df_anonimo))
    else:
        riepilogo.to_csv(args.out, index=False, encoding="utf-8")
        print("Nota: il .csv contiene solo il Riepilogo. Per il dettaglio e il traffico anonimo usa --out report.xlsx.")

    print(f"Fatto. Report salvato in: {args.out}")
    print(f"Contatti/ref trovati: {len(profile)}")


if __name__ == "__main__":
    main()
