"""
Genera/aggiorna l'array `studi` dentro generatore_email.html a partire
dalla lista contatti — CSV o XLSX (rilevato dall'estensione, come negli
altri due script della cartella).

Il generatore non puo' leggere il file contatti da solo a runtime
(aperto via file://, il browser blocca fetch locali per CORS), quindi
l'array va rigenerato con questo script ogni volta che la lista contatti
cambia — cosi' generatore e lista non possono piu' divergere.

Il link di ogni contatto viene costruito da --dominio + ref: la colonna
'link' della lista viene ignorata (puo' puntare al dominio vecchio o a
quello di un altro pubblico). Le colonne citta/categoria, se assenti
nella lista, vengono preservate dai valori gia' presenti nel generatore,
per ref.

XLSX CON PIU' FOGLI (es. Lista_Contatti_Rilievo_Contract.xlsx, tab
Architetti + Agenzie immobiliari): il generatore e' sempre specifico per
UN pubblico (un dominio solo), quindi va scelto ESPLICITAMENTE quale
foglio usare con --foglio. Se il file ha un solo foglio dati (escluso
uno chiamato "Riepilogo"), --foglio si puo' omettere: lo sceglie da
solo. Con piu' fogli dati e --foglio omesso, lo script si ferma ed
elenca i fogli disponibili — non sceglie mai al posto tuo, altrimenti
si rischia di mischiare due pubblici nello stesso generatore.

ATTENZIONE MULTI-PUBBLICO (dominio): ogni branch/pubblico ha il proprio
hostname (es. collaboratori.rilievocontract.it per gli architetti,
agenzie-immobiliari.rilievocontract.it per le agenzie). --dominio non ha
un default fisso proprio per questo: uno sbagliato riscrive
silenziosamente TUTTI i link del generatore su un sito che non e' il
suo (bug reale, trovato il 20/07/2026 lanciando lo script sul
generatore agenzie con il default hardcoded degli architetti — tutti i
19 link erano finiti su collaboratori.rilievocontract.it invece che su
agenzie-immobiliari.rilievocontract.it). Il parametro e' quindi
obbligatorio: lo script si rifiuta di partire senza, niente da
"indovinare" per sbaglio.

USO:
    python tools/genera_generatore.py lista.csv --dominio "https://collaboratori.rilievocontract.it/?ref="
    python tools/genera_generatore.py Lista_Contatti_Rilievo_Contract.xlsx --foglio "Agenzie immobiliari" --dominio "https://agenzie-immobiliari.rilievocontract.it/?ref=" --generatore generatore_email.html
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_DISPONIBILE = True
except ImportError:
    OPENPYXL_DISPONIBILE = False

# Il generatore sta nella root del progetto, questo script in tools/ga4-report/
PERCORSO_GENERATORE_DEFAULT = Path(__file__).resolve().parent.parent.parent / "generatore_email.html"

FOGLIO_DA_IGNORARE = "riepilogo"

ALIAS_COLONNE = {
    "id": "id", "nome": "nome", "città": "citta", "citta": "citta",
    "email": "email", "categoria": "categoria", "ref": "ref",
    "link": "link", "link presentazione": "link", "tipo": "tipo",
}


def _canonicalizza(testo):
    chiave = (testo or "").strip().lower()
    return ALIAS_COLONNE.get(chiave, chiave)


def leggi_contatti_csv(percorso):
    with open(percorso, newline="", encoding="utf-8-sig") as f:
        righe = list(csv.DictReader(f))
    if not righe:
        sys.exit("Il CSV contatti e' vuoto.")
    return righe


def leggi_contatti_xlsx(percorso, nome_foglio_scelto):
    if not OPENPYXL_DISPONIBILE:
        sys.exit("Serve 'openpyxl' per leggere file .xlsx: pip install openpyxl --break-system-packages")

    wb = load_workbook(percorso, data_only=True)
    fogli_dati = [n for n in wb.sheetnames if n.strip().lower() != FOGLIO_DA_IGNORARE]
    if not fogli_dati:
        sys.exit(f"Nessun foglio dati in {percorso}.")

    if nome_foglio_scelto:
        if nome_foglio_scelto not in wb.sheetnames:
            sys.exit(f"Foglio '{nome_foglio_scelto}' non trovato. Fogli disponibili: {', '.join(wb.sheetnames)}")
        foglio = nome_foglio_scelto
    elif len(fogli_dati) == 1:
        foglio = fogli_dati[0]
    else:
        sys.exit(
            f"{percorso} ha piu' fogli dati ({', '.join(fogli_dati)}) — specifica quale usare con "
            f"--foglio, es. --foglio \"{fogli_dati[0]}\". Il generatore serve un pubblico solo per volta."
        )

    ws = wb[foglio]
    righe_grezze = list(ws.iter_rows(values_only=True))
    if not righe_grezze:
        sys.exit(f"Il foglio '{foglio}' e' vuoto.")

    intestazione = [_canonicalizza(c) for c in righe_grezze[0]]
    righe = []
    for riga_valori in righe_grezze[1:]:
        if all(v is None or str(v).strip() == "" for v in riga_valori):
            continue
        riga = {}
        for i, nome_colonna in enumerate(intestazione):
            if not nome_colonna:
                continue
            valore = riga_valori[i] if i < len(riga_valori) else None
            riga[nome_colonna] = "" if valore is None else str(valore).strip()
        righe.append(riga)

    print(f"Foglio usato: '{foglio}' ({len(righe)} contatti).")
    return righe


def leggi_contatti(percorso, nome_foglio_scelto):
    """Legge la lista contatti da CSV o XLSX. Colonne richieste: nome, email, ref."""
    estensione = Path(percorso).suffix.lower()
    if estensione == ".csv":
        righe = leggi_contatti_csv(percorso)
    elif estensione in (".xlsx", ".xlsm"):
        righe = leggi_contatti_xlsx(percorso, nome_foglio_scelto)
    else:
        sys.exit(f"Formato non supportato: '{estensione}' (atteso .csv o .xlsx).")

    colonne_richieste = ["nome", "email", "ref"]
    for colonna in colonne_richieste:
        if colonna not in righe[0]:
            sys.exit(f"Colonna '{colonna}' mancante nella lista contatti (richieste: {colonne_richieste}).")

    ref_visti = set()
    for riga in righe:
        riga["ref"] = riga["ref"].strip()
        if riga["ref"] in ref_visti:
            sys.exit(f"Ref duplicato nella lista: {riga['ref']} — correggerla prima di rigenerare.")
        ref_visti.add(riga["ref"])

    return righe


def estrai_studi_esistenti(html):
    """Estrae l'array `studi` attuale dal generatore, per preservare
    citta/categoria (che non esistono nel CSV)."""
    match = re.search(r"const studi = (\[.*?\]);", html, re.DOTALL)
    if match is None:
        sys.exit("Non trovo 'const studi = [...];' nel generatore: file inatteso.")
    return json.loads(match.group(1)), match


def main():
    parser = argparse.ArgumentParser(description="Rigenera l'array studi del generatore email dalla lista contatti (CSV o XLSX)")
    parser.add_argument("file_contatti", help="CSV o XLSX con colonne id,nome,email,ref (la colonna link e' ignorata)")
    parser.add_argument("--foglio", help="Nome del foglio da usare se il file e' un XLSX con piu' fogli dati "
                                          "(es. \"Agenzie immobiliari\"). Non serve se il file ha un solo foglio dati.")
    parser.add_argument("--generatore", default=str(PERCORSO_GENERATORE_DEFAULT),
                        help="Percorso di generatore_email.html (default: root del progetto)")
    parser.add_argument("--dominio", required=True,
                        help="Dominio del pubblico, con ?ref= incluso, es. "
                             "'https://agenzie-immobiliari.rilievocontract.it/?ref='. "
                             "Obbligatorio, niente default: un dominio sbagliato riscrive "
                             "in silenzio tutti i link sul sito sbagliato.")
    args = parser.parse_args()

    if not Path(args.file_contatti).exists():
        sys.exit(f"File non trovato: {args.file_contatti}")
    if not Path(args.generatore).exists():
        sys.exit(f"Generatore non trovato: {args.generatore}")

    contatti = leggi_contatti(args.file_contatti, args.foglio)

    html = Path(args.generatore).read_text(encoding="utf-8")
    studi_esistenti, match = estrai_studi_esistenti(html)

    # citta/categoria per ref, dai dati gia' presenti nel generatore
    extra_per_ref = {}
    for studio in studi_esistenti:
        extra_per_ref[studio.get("ref", "")] = {
            "citta": studio.get("citta", ""),
            "categoria": studio.get("categoria", ""),
        }

    # citta/categoria: preferisci quelle della lista contatti (fonte di
    # verita' aggiornata, es. Lista_Contatti_Rilievo_Contract.xlsx le ha
    # gia' compilate) — quelle preservate dal generatore esistente sono
    # solo un fallback per liste piu' vecchie/minime che non le hanno
    # (es. un CSV con solo nome,email,ref). Prima di questa correzione
    # (20/07/2026) il generatore ignorava sempre citta/categoria della
    # lista, anche quando erano presenti: bug trovato rigenerando il
    # generatore agenzie con la lista che le aveva gia' tutte compilate.
    nuovi_studi = []
    da_lista = 0
    da_generatore = 0
    for contatto in contatti:
        extra_generatore = extra_per_ref.get(contatto["ref"], {"citta": "", "categoria": ""})
        citta = contatto.get("citta", "").strip() or extra_generatore["citta"]
        categoria = contatto.get("categoria", "").strip() or extra_generatore["categoria"]
        if contatto.get("citta", "").strip() or contatto.get("categoria", "").strip():
            da_lista += 1
        elif contatto["ref"] in extra_per_ref:
            da_generatore += 1
        nuovi_studi.append({
            "id": contatto.get("id", "").strip(),
            "nome": contatto["nome"].strip(),
            "citta": citta,
            "email": contatto["email"].strip(),
            "categoria": categoria,
            "ref": contatto["ref"],
            "link": args.dominio + contatto["ref"],
        })

    nuova_riga = "const studi = " + json.dumps(nuovi_studi, ensure_ascii=False) + ";"
    html_nuovo = html[: match.start()] + nuova_riga + html[match.end():]
    Path(args.generatore).write_text(html_nuovo, encoding="utf-8")

    print(f"Fatto. {len(nuovi_studi)} studi scritti in: {args.generatore}")
    print(f"citta/categoria dalla lista contatti: {da_lista} | ereditate dal generatore esistente: {da_generatore} | "
          f"senza nessuna delle due: {len(nuovi_studi) - da_lista - da_generatore}")


if __name__ == "__main__":
    main()
